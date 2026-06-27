import logging
from datetime import datetime
from html import escape
from pathlib import Path

import openpyxl
from config import REPORTS_DIR
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

NAVY = "1F3864"
WHITE = "FFFFFF"
RED_FILL = "FFCCCC"
YELLOW_FILL = "FFF2CC"
GREEN_FILL = "CCFFCC"
GREY_FILL = "F2F2F2"
ALT_FILL = "EEF2FF"
ORANGE_FILL = "FFE5CC"
MAX_HTML_PICKLIST_ROWS = 500


def _header_style(ws, row: int, cols: list[str]):
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(bold=True, color=WHITE)
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _row_fill(ws, row: int, ncols: int, color: str):
    fill = PatternFill("solid", fgColor=color)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def generate_excel_report(
    alias_a: str, alias_b: str, result: dict, instance_a: dict, instance_b: dict
) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = REPORTS_DIR / f"{alias_a}_vs_{alias_b}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    _write_summary_sheet(ws, alias_a, alias_b, result, instance_a, instance_b, timestamp)

    ws_e = wb.create_sheet("Missing Entities")
    _write_entity_diff_sheet(ws_e, result["entity_diffs"], alias_a, alias_b)

    missing_fields = [d for d in result["field_diffs"] if "only" in d["diff_type"]]
    attr_diffs = [d for d in result["field_diffs"] if d["diff_type"] == "attribute_mismatch"]

    ws_mf = wb.create_sheet("Missing Fields")
    _write_fields_only_sheet(ws_mf, missing_fields, alias_a, alias_b)

    ws_fd = wb.create_sheet("Field Attr Diffs")
    _write_field_diff_sheet(ws_fd, attr_diffs, alias_a, alias_b)

    pr = result["picklist_result"]

    ws_mp = wb.create_sheet("Missing Picklists")
    _write_missing_picklists_sheet(ws_mp, pr["missing_picklists"], alias_a, alias_b)

    ws_mv = wb.create_sheet("Missing Values")
    _write_missing_values_sheet(ws_mv, pr["missing_values"], alias_a, alias_b)

    ws_vd = wb.create_sheet("Value Diffs")
    _write_value_diffs_sheet(ws_vd, pr["value_diffs"], alias_a, alias_b)

    wb.save(str(fname))
    logger.info("Excel report saved: %s", fname)
    return fname


def _write_summary_sheet(ws, alias_a, alias_b, result, inst_a, inst_b, ts):
    s = result["summary"]
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 25

    fill_h = PatternFill("solid", fgColor=NAVY)
    font_h = Font(bold=True, color=WHITE, size=14)
    title_cell = ws.cell(row=1, column=1, value="SF Config Compare - Comparison Report")
    title_cell.fill = fill_h
    title_cell.font = font_h
    ws.merge_cells("A1:B1")

    data = [
        ("Generated", ts),
        ("Instance A", f"{alias_a} ({inst_a.get('base_url', '')})"),
        ("Instance B", f"{alias_b} ({inst_b.get('base_url', '')})"),
        ("", ""),
        ("Entities only in A", s["entities_only_in_a"]),
        ("Entities only in B", s["entities_only_in_b"]),
        ("Entities in both", s["entities_in_both"]),
        ("", ""),
        ("Fields matched", s["fields_matched"]),
        ("Fields with differences", s["fields_with_diff"]),
        ("Fields only in A", s["fields_only_in_a"]),
        ("Fields only in B", s["fields_only_in_b"]),
        ("", ""),
        ("Picklists only in A (entire)", s.get("picklists_only_in_a", 0)),
        ("Picklists only in B (entire)", s.get("picklists_only_in_b", 0)),
        ("Picklists shared (compared)", s.get("picklists_shared", 0)),
        ("Values missing from A", s.get("missing_values_in_a", 0)),
        ("Values missing from B", s.get("missing_values_in_b", 0)),
        ("Value field differences", s.get("value_diffs", 0)),
    ]
    for r, (label, value) in enumerate(data, 3):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)

    ws.freeze_panes = "A2"


def _write_entity_diff_sheet(ws, entity_diffs, alias_a, alias_b):
    cols = ["Entity Name", "Diff Type", "Entity Label", "Element Name"]
    _header_style(ws, 1, cols)
    ws.freeze_panes = "A2"
    for i, d in enumerate(entity_diffs, 2):
        details = d.get("details") or {}
        ws.cell(row=i, column=1, value=d["entity_name"])
        ws.cell(row=i, column=2, value=d["diff_type"].replace("_", " ").title())
        ws.cell(row=i, column=3, value=details.get("entity_label"))
        ws.cell(row=i, column=4, value=details.get("element_name"))
        _row_fill(ws, i, len(cols), RED_FILL if "only" in d["diff_type"] else ALT_FILL)
    _auto_width(ws)


def _write_field_diff_sheet(ws, field_diffs, alias_a, alias_b):
    cols = [
        "Entity",
        "Field ID",
        "Field Label",
        "Attribute",
        f"Value in {alias_a}",
        f"Value in {alias_b}",
        "Diff Type",
    ]
    _header_style(ws, 1, cols)
    ws.freeze_panes = "A2"
    for i, d in enumerate(field_diffs, 2):
        ws.cell(row=i, column=1, value=d["entity_name"])
        ws.cell(row=i, column=2, value=d["field_id"])
        ws.cell(row=i, column=3, value=d.get("field_label"))
        ws.cell(row=i, column=4, value=d.get("attribute"))
        ws.cell(row=i, column=5, value=d.get("value_a"))
        ws.cell(row=i, column=6, value=d.get("value_b"))
        ws.cell(row=i, column=7, value=d["diff_type"].replace("_", " ").title())
        if "only" in d["diff_type"]:
            _row_fill(ws, i, len(cols), RED_FILL)
        else:
            _row_fill(ws, i, len(cols), YELLOW_FILL if i % 2 == 0 else "FFFDE7")
    _auto_width(ws)


def _write_missing_picklists_sheet(ws, missing_picklists, alias_a, alias_b):
    cols = ["Picklist ID", "Missing From", "Value Count in Other Instance"]
    _header_style(ws, 1, cols)
    ws.freeze_panes = "A2"
    for i, d in enumerate(missing_picklists, 2):
        missing_from = alias_b if d["diff_type"] == "only_in_a" else alias_a
        ws.cell(row=i, column=1, value=d["picklist_id"])
        ws.cell(row=i, column=2, value=missing_from)
        ws.cell(row=i, column=3, value=d["value_count"])
        _row_fill(ws, i, len(cols), RED_FILL if d["diff_type"] == "only_in_a" else ALT_FILL)
    _auto_width(ws)


def _write_missing_values_sheet(ws, missing_values, alias_a, alias_b):
    cols = ["Picklist ID", "Option Code", "Label", "Status", "Missing From"]
    _header_style(ws, 1, cols)
    ws.freeze_panes = "A2"
    for i, d in enumerate(missing_values, 2):
        missing_from = alias_b if d["diff_type"] == "only_in_a" else alias_a
        ws.cell(row=i, column=1, value=d["picklist_id"])
        ws.cell(row=i, column=2, value=d["external_code"])
        ws.cell(row=i, column=3, value=d.get("label"))
        ws.cell(row=i, column=4, value=d.get("status"))
        ws.cell(row=i, column=5, value=missing_from)
        _row_fill(ws, i, len(cols), RED_FILL if d["diff_type"] == "only_in_a" else ALT_FILL)
    _auto_width(ws)


def _write_value_diffs_sheet(ws, value_diffs, alias_a, alias_b):
    cols = [
        "Picklist ID",
        "Option Code",
        f"Label in {alias_a}",
        f"Label in {alias_b}",
        "Field",
        f"Value in {alias_a}",
        f"Value in {alias_b}",
    ]
    _header_style(ws, 1, cols)
    ws.freeze_panes = "A2"
    row = 2
    for d in value_diffs:
        for fd in d["field_diffs"]:
            ws.cell(row=row, column=1, value=d["picklist_id"])
            ws.cell(row=row, column=2, value=d["external_code"])
            ws.cell(row=row, column=3, value=d.get("label_a"))
            ws.cell(row=row, column=4, value=d.get("label_b"))
            ws.cell(row=row, column=5, value=fd["field"])
            ws.cell(row=row, column=6, value=fd["value_a"])
            ws.cell(row=row, column=7, value=fd["value_b"])
            _row_fill(ws, row, len(cols), YELLOW_FILL if row % 2 == 0 else "FFFDE7")
            row += 1
    _auto_width(ws)


def _write_fields_only_sheet(ws, field_diffs, alias_a, alias_b):
    cols = ["Entity", "Field ID", "Field Label", "Missing From"]
    _header_style(ws, 1, cols)
    ws.freeze_panes = "A2"
    for i, d in enumerate(field_diffs, 2):
        missing_from = alias_b if d["diff_type"] == "only_in_a" else alias_a
        ws.cell(row=i, column=1, value=d["entity_name"])
        ws.cell(row=i, column=2, value=d["field_id"])
        ws.cell(row=i, column=3, value=d.get("field_label"))
        ws.cell(row=i, column=4, value=missing_from)
        _row_fill(ws, i, len(cols), RED_FILL if d["diff_type"] == "only_in_a" else ALT_FILL)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------


def generate_html_report(
    alias_a: str,
    alias_b: str,
    result: dict,
    download_url: str | None = None,
    nav_urls: dict[str, str] | None = None,
) -> str:
    s = result["summary"]
    entity_diffs = result["entity_diffs"]
    field_diffs = result["field_diffs"]
    pr = result["picklist_result"]
    missing_picklists = pr["missing_picklists"]
    missing_values = pr["missing_values"]
    value_diffs = pr["value_diffs"]

    def h(value):
        return escape(str(value or ""))

    # --- Summary cards ---
    nav_urls = nav_urls or {}
    action_links = []
    if nav_urls.get("dashboard"):
        action_links.append(
            f'<a href="{h(nav_urls["dashboard"])}" class="inline-flex items-center rounded-lg border border-gray-300 bg-white px-4 py-2 text-sm font-semibold text-gray-700 hover:bg-gray-50">Back to Dashboard</a>'
        )
    if nav_urls.get("compare"):
        action_links.append(
            f'<a href="{h(nav_urls["compare"])}" class="inline-flex items-center rounded-lg border border-gray-300 bg-white px-4 py-2 text-sm font-semibold text-gray-700 hover:bg-gray-50">Run Another Compare</a>'
        )
    if download_url:
        action_links.append(
            f'<a href="{h(download_url)}" class="inline-flex items-center rounded-lg bg-indigo-600 px-4 py-2 text-sm font-semibold text-white hover:bg-indigo-700">Download Excel Report</a>'
        )
    report_actions = ""
    if action_links:
        report_actions = f'<div class="flex flex-wrap gap-2">{"".join(action_links)}</div>'

    total_picklist_issues = len(missing_picklists) + len(missing_values) + len(value_diffs)
    summary_cards = f"""
    <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mt-4">
        <div class="bg-gray-50 rounded-lg p-4 text-center"><div class="text-2xl font-bold">{s["entities_in_both"]}</div><div class="text-xs text-gray-500 mt-1">Entities Compared</div></div>
        <div class="{"bg-red-50" if s["fields_with_diff"] + s["fields_only_in_a"] + s["fields_only_in_b"] else "bg-green-50"} rounded-lg p-4 text-center">
            <div class="text-2xl font-bold {"text-red-700" if s["fields_with_diff"] + s["fields_only_in_a"] + s["fields_only_in_b"] else "text-green-700"}">{s["fields_with_diff"] + s["fields_only_in_a"] + s["fields_only_in_b"]}</div>
            <div class="text-xs text-gray-500 mt-1">Field Diffs</div></div>
        <div class="{"bg-amber-50" if total_picklist_issues else "bg-green-50"} rounded-lg p-4 text-center">
            <div class="text-2xl font-bold {"text-amber-700" if total_picklist_issues else "text-green-700"}">{total_picklist_issues}</div>
            <div class="text-xs text-gray-500 mt-1">Picklist Issues</div></div>
        <div class="bg-green-50 rounded-lg p-4 text-center"><div class="text-2xl font-bold text-green-700">{s["fields_matched"]}</div><div class="text-xs text-gray-500 mt-1">Fields Matched</div></div>
    </div>"""

    # Split field_diffs into missing fields and attribute mismatches
    missing_fields = [d for d in field_diffs if "only" in d["diff_type"]]
    attr_diffs = [d for d in field_diffs if d["diff_type"] == "attribute_mismatch"]

    # --- Missing entities panel ---
    entity_rows = ""
    for d in entity_diffs:
        missing_from = alias_b if d["diff_type"] == "only_in_a" else alias_a
        badge_cls = (
            "bg-red-100 text-red-800"
            if d["diff_type"] == "only_in_a"
            else "bg-blue-100 text-blue-800"
        )
        entity_rows += f"""<tr class="border-b hover:bg-gray-50">
            <td class="px-4 py-2 font-mono text-sm">{h(d["entity_name"])}</td>
            <td class="px-4 py-2"><span class="px-2 py-0.5 rounded text-xs font-medium {badge_cls}">Missing from {h(missing_from)}</span></td>
        </tr>"""

    # --- Missing fields panel (grouped by entity) ---
    mf_by_entity: dict[str, list] = {}
    for d in missing_fields:
        mf_by_entity.setdefault(d["entity_name"], []).append(d)

    mf_sections = ""
    for entity_name, diffs in sorted(mf_by_entity.items()):
        rows = ""
        for d in diffs:
            missing_from = alias_b if d["diff_type"] == "only_in_a" else alias_a
            badge_cls = (
                "bg-red-100 text-red-800"
                if d["diff_type"] == "only_in_a"
                else "bg-blue-100 text-blue-800"
            )
            rows += f"""<tr class="border-b hover:bg-gray-50">
                <td class="px-3 py-1.5 font-mono text-xs">{h(d["field_id"])}</td>
                <td class="px-3 py-1.5 text-sm">{h(d.get("field_label"))}</td>
                <td class="px-3 py-1.5"><span class="px-2 py-0.5 rounded text-xs font-medium {badge_cls}">Missing from {h(missing_from)}</span></td>
            </tr>"""
        mf_sections += f"""
        <details class="mb-2 border rounded-lg overflow-hidden">
            <summary class="bg-gray-50 px-4 py-2 cursor-pointer font-semibold text-sm hover:bg-gray-100 flex items-center justify-between">
                <span class="font-mono">{h(entity_name)}</span>
                <span class="text-xs font-normal text-gray-500 ml-2">{len(diffs)} missing field{"s" if len(diffs) != 1 else ""}</span>
            </summary>
            <table class="w-full text-sm">
                <thead><tr class="bg-gray-50 text-xs text-gray-500 uppercase">
                    <th class="px-3 py-2 text-left">Field ID</th>
                    <th class="px-3 py-2 text-left">Label</th>
                    <th class="px-3 py-2 text-left">Where</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </details>"""

    # --- Attribute diffs panel (grouped by entity) ---
    ad_by_entity: dict[str, list] = {}
    for d in attr_diffs:
        ad_by_entity.setdefault(d["entity_name"], []).append(d)

    ad_sections = ""
    for entity_name, diffs in sorted(ad_by_entity.items()):
        rows = ""
        for d in diffs:
            field_chips = f"""<div class="flex items-center gap-2 text-xs">
                <span class="font-mono text-gray-500 w-24 shrink-0">{h(d.get("attribute"))}</span>
                <span class="text-red-700 bg-red-50 rounded px-1.5 py-0.5 max-w-xs truncate" title="{h(d.get("value_a"))}">{h(d.get("value_a")) or '<em class="text-gray-400">empty</em>'}</span>
                <span class="text-gray-400">→</span>
                <span class="text-blue-700 bg-blue-50 rounded px-1.5 py-0.5 max-w-xs truncate" title="{h(d.get("value_b"))}">{h(d.get("value_b")) or '<em class="text-gray-400">empty</em>'}</span>
            </div>"""
            rows += f"""<tr class="border-b hover:bg-gray-50">
                <td class="px-3 py-2 font-mono text-xs align-top">{h(d["field_id"])}</td>
                <td class="px-3 py-2 text-sm align-top">{h(d.get("field_label"))}</td>
                <td class="px-3 py-2 align-top">{field_chips}</td>
            </tr>"""
        ad_sections += f"""
        <details class="mb-2 border rounded-lg overflow-hidden">
            <summary class="bg-gray-50 px-4 py-2 cursor-pointer font-semibold text-sm hover:bg-gray-100 flex items-center justify-between">
                <span class="font-mono">{h(entity_name)}</span>
                <span class="text-xs font-normal text-gray-500 ml-2">{len(diffs)} diff{"s" if len(diffs) != 1 else ""}</span>
            </summary>
            <table class="w-full text-sm">
                <thead><tr class="bg-gray-50 text-xs text-gray-500 uppercase">
                    <th class="px-3 py-2 text-left">Field ID</th>
                    <th class="px-3 py-2 text-left">Label</th>
                    <th class="px-3 py-2 text-left">Attribute Differences</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </details>"""

    # --- Missing picklists section ---
    mp_rows = ""
    for d in missing_picklists[:MAX_HTML_PICKLIST_ROWS]:
        missing_from = alias_b if d["diff_type"] == "only_in_a" else alias_a
        badge_cls = (
            "bg-red-100 text-red-800"
            if d["diff_type"] == "only_in_a"
            else "bg-blue-100 text-blue-800"
        )
        mp_rows += f"""<tr class="border-b hover:bg-gray-50">
            <td class="px-4 py-2 font-mono text-sm">{h(d["picklist_id"])}</td>
            <td class="px-4 py-2"><span class="px-2 py-0.5 rounded text-xs font-medium {badge_cls}">Missing from {h(missing_from)}</span></td>
            <td class="px-4 py-2 text-sm text-gray-600">{d["value_count"]} values</td>
        </tr>"""

    # --- Missing values section (grouped by picklist) ---
    mv_by_pl: dict[str, list] = {}
    for v in missing_values:
        mv_by_pl.setdefault(v["picklist_id"], []).append(v)

    mv_sections = ""
    shown_mv = 0
    for pl_id, vals in sorted(mv_by_pl.items()):
        if shown_mv >= MAX_HTML_PICKLIST_ROWS:
            break
        rows = ""
        for v in vals:
            if shown_mv >= MAX_HTML_PICKLIST_ROWS:
                break
            missing_from = alias_b if v["diff_type"] == "only_in_a" else alias_a
            badge_cls = (
                "bg-red-100 text-red-800"
                if v["diff_type"] == "only_in_a"
                else "bg-blue-100 text-blue-800"
            )
            rows += f"""<tr class="border-b hover:bg-gray-50">
                <td class="px-3 py-1.5 font-mono text-xs">{h(v["external_code"])}</td>
                <td class="px-3 py-1.5 text-sm">{h(v.get("label"))}</td>
                <td class="px-3 py-1.5 text-sm">{h(v.get("status"))}</td>
                <td class="px-3 py-1.5"><span class="px-2 py-0.5 rounded text-xs font-medium {badge_cls}">Missing from {h(missing_from)}</span></td>
            </tr>"""
            shown_mv += 1
        mv_sections += f"""
        <details class="mb-2 border rounded-lg overflow-hidden">
            <summary class="bg-gray-50 px-4 py-2 cursor-pointer font-semibold text-sm hover:bg-gray-100 flex items-center justify-between">
                <span class="font-mono">{h(pl_id)}</span>
                <span class="text-xs font-normal text-gray-500 ml-2">{len(vals)} missing value{"s" if len(vals) != 1 else ""}</span>
            </summary>
            <table class="w-full text-sm">
                <thead><tr class="bg-gray-50 text-xs text-gray-500 uppercase">
                    <th class="px-3 py-2 text-left">Option Code</th><th class="px-3 py-2 text-left">Label</th>
                    <th class="px-3 py-2 text-left">Status</th><th class="px-3 py-2 text-left">Where</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </details>"""

    # --- Value diffs section (grouped by picklist) ---
    vd_by_pl: dict[str, list] = {}
    for v in value_diffs:
        vd_by_pl.setdefault(v["picklist_id"], []).append(v)

    vd_sections = ""
    shown_vd = 0
    for pl_id, vals in sorted(vd_by_pl.items()):
        if shown_vd >= MAX_HTML_PICKLIST_ROWS:
            break
        rows = ""
        for v in vals:
            if shown_vd >= MAX_HTML_PICKLIST_ROWS:
                break
            field_chips = ""
            for fd in v["field_diffs"]:
                field_chips += f"""<div class="flex items-center gap-2 text-xs py-0.5">
                    <span class="font-mono text-gray-500 w-28 shrink-0">{h(fd["field"])}</span>
                    <span class="text-red-700 bg-red-50 rounded px-1.5 py-0.5 max-w-xs truncate" title="{h(fd["value_a"])}">{h(fd["value_a"]) or '<em class="text-gray-400">empty</em>'}</span>
                    <span class="text-gray-400">→</span>
                    <span class="text-blue-700 bg-blue-50 rounded px-1.5 py-0.5 max-w-xs truncate" title="{h(fd["value_b"])}">{h(fd["value_b"]) or '<em class="text-gray-400">empty</em>'}</span>
                </div>"""
            rows += f"""<tr class="border-b hover:bg-gray-50">
                <td class="px-3 py-2 font-mono text-xs align-top">{h(v["external_code"])}</td>
                <td class="px-3 py-2 text-sm text-red-700 align-top">{h(v.get("label_a"))}</td>
                <td class="px-3 py-2 text-sm text-blue-700 align-top">{h(v.get("label_b"))}</td>
                <td class="px-3 py-2 align-top">{field_chips}</td>
            </tr>"""
            shown_vd += 1
        vd_sections += f"""
        <details class="mb-2 border rounded-lg overflow-hidden">
            <summary class="bg-gray-50 px-4 py-2 cursor-pointer font-semibold text-sm hover:bg-gray-100 flex items-center justify-between">
                <span class="font-mono">{h(pl_id)}</span>
                <span class="text-xs font-normal text-gray-500 ml-2">{len(vals)} value diff{"s" if len(vals) != 1 else ""}</span>
            </summary>
            <table class="w-full text-sm">
                <thead><tr class="bg-gray-50 text-xs text-gray-500 uppercase">
                    <th class="px-3 py-2 text-left">Option Code</th>
                    <th class="px-3 py-2 text-left">Label in {h(alias_a)}</th>
                    <th class="px-3 py-2 text-left">Label in {h(alias_b)}</th>
                    <th class="px-3 py-2 text-left">Field Differences</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>
        </details>"""

    overflow_note = ""
    total_shown = shown_mv + shown_vd
    if total_shown >= MAX_HTML_PICKLIST_ROWS:
        overflow_note = f'<p class="mb-3 text-sm text-amber-700 bg-amber-50 border border-amber-200 rounded-lg px-3 py-2">HTML preview capped at {MAX_HTML_PICKLIST_ROWS} rows per section. Download Excel for complete data.</p>'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Diff Report: {h(alias_a)} vs {h(alias_b)}</title>
<script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-gray-50 font-sans">
<div class="max-w-7xl mx-auto px-4 py-8">

  <!-- Header -->
  <div class="bg-white rounded-xl shadow p-6 mb-6">
    <div class="flex flex-col gap-3 md:flex-row md:items-start md:justify-between">
      <div>
        <h1 class="text-2xl font-bold text-gray-900 mb-1">Comparison Report</h1>
        <p class="text-gray-500 text-sm">
          <span class="font-semibold text-red-700">{h(alias_a)}</span>
          <span class="mx-2 text-gray-400">vs</span>
          <span class="font-semibold text-blue-700">{h(alias_b)}</span>
        </p>
      </div>
      {report_actions}
    </div>
    {summary_cards}
  </div>

  <!-- Filter -->
  <div class="mb-4">
    <input type="text" id="filter" placeholder="Filter by picklist ID, option code, field…"
           oninput="filterContent(this.value)"
           class="w-full border rounded-lg px-4 py-2 text-sm focus:outline-none focus:ring-2 focus:ring-indigo-500">
  </div>

  <!-- Metadata Section -->
  <div class="bg-white rounded-xl shadow p-6 mb-6" id="metadata-section">
    <h2 class="text-lg font-semibold mb-1">Metadata Differences</h2>
    <div class="flex gap-4 text-sm text-gray-500 mb-4">
      <span class="bg-red-50 text-red-700 rounded px-2 py-0.5 font-medium">{len(entity_diffs)} missing entities</span>
      <span class="bg-orange-50 text-orange-700 rounded px-2 py-0.5 font-medium">{len(missing_fields)} missing fields</span>
      <span class="bg-yellow-50 text-yellow-700 rounded px-2 py-0.5 font-medium">{len(attr_diffs)} attribute diffs</span>
    </div>

    {'<details class="mb-4 border border-red-200 rounded-xl overflow-hidden" open><summary class="bg-red-50 px-4 py-3 cursor-pointer font-semibold text-sm text-red-800 hover:bg-red-100 flex items-center justify-between"><span>Entire Entities Missing</span><span class="text-xs font-normal">' + str(len(entity_diffs)) + ' entities</span></summary><table class="w-full text-sm"><thead><tr class="bg-gray-50 text-xs text-gray-500 uppercase"><th class="px-4 py-2 text-left">Entity Name</th><th class="px-4 py-2 text-left">Missing From</th></tr></thead><tbody>' + entity_rows + "</tbody></table></details>" if entity_diffs else '<p class="text-sm text-green-700 bg-green-50 rounded-lg px-3 py-2 mb-4">No missing entities.</p>'}

    {'<details class="mb-4 border border-orange-200 rounded-xl overflow-hidden"><summary class="bg-orange-50 px-4 py-3 cursor-pointer font-semibold text-sm text-orange-800 hover:bg-orange-100 flex items-center justify-between"><span>Missing Fields within Shared Entities</span><span class="text-xs font-normal">' + str(len(missing_fields)) + " fields across " + str(len(mf_by_entity)) + ' entities</span></summary><div class="p-4">' + mf_sections + "</div></details>" if missing_fields else '<p class="text-sm text-green-700 bg-green-50 rounded-lg px-3 py-2 mb-4">No missing fields within shared entities.</p>'}

    {'<details class="mb-4 border border-yellow-200 rounded-xl overflow-hidden"><summary class="bg-yellow-50 px-4 py-3 cursor-pointer font-semibold text-sm text-yellow-800 hover:bg-yellow-100 flex items-center justify-between"><span>Field Attribute Differences</span><span class="text-xs font-normal">' + str(len(attr_diffs)) + " fields across " + str(len(ad_by_entity)) + ' entities</span></summary><div class="p-4">' + ad_sections + "</div></details>" if attr_diffs else '<p class="text-sm text-green-700 bg-green-50 rounded-lg px-3 py-2 mb-4">No field attribute differences.</p>'}
  </div>

  <!-- Picklist Section -->
  <div class="bg-white rounded-xl shadow p-6 mb-6" id="picklist-section">
    <h2 class="text-lg font-semibold mb-1">Picklist Differences</h2>
    <div class="flex gap-4 text-sm text-gray-500 mb-4">
      <span class="bg-red-50 text-red-700 rounded px-2 py-0.5 font-medium">{len(missing_picklists)} missing picklists</span>
      <span class="bg-orange-50 text-orange-700 rounded px-2 py-0.5 font-medium">{len(missing_values)} missing values</span>
      <span class="bg-yellow-50 text-yellow-700 rounded px-2 py-0.5 font-medium">{len(value_diffs)} value diffs</span>
    </div>
    {overflow_note}

    <!-- Missing Picklists -->
    {'<details class="mb-4 border border-red-200 rounded-xl overflow-hidden" open><summary class="bg-red-50 px-4 py-3 cursor-pointer font-semibold text-sm text-red-800 hover:bg-red-100 flex items-center justify-between"><span>Entire Picklists Missing</span><span class="text-xs font-normal">' + str(len(missing_picklists)) + ' picklists</span></summary><table class="w-full text-sm"><thead><tr class="bg-gray-50 text-xs text-gray-500 uppercase"><th class="px-4 py-2 text-left">Picklist ID</th><th class="px-4 py-2 text-left">Missing From</th><th class="px-4 py-2 text-left">Values in Other</th></tr></thead><tbody>' + mp_rows + "</tbody></table></details>" if missing_picklists else '<p class="text-sm text-green-700 bg-green-50 rounded-lg px-3 py-2 mb-4">No entire picklists missing.</p>'}

    <!-- Missing Values -->
    {'<details class="mb-4 border border-orange-200 rounded-xl overflow-hidden"><summary class="bg-orange-50 px-4 py-3 cursor-pointer font-semibold text-sm text-orange-800 hover:bg-orange-100 flex items-center justify-between"><span>Missing Values within Shared Picklists</span><span class="text-xs font-normal">' + str(len(missing_values)) + " values across " + str(len(mv_by_pl)) + ' picklists</span></summary><div class="p-4">' + mv_sections + "</div></details>" if missing_values else '<p class="text-sm text-green-700 bg-green-50 rounded-lg px-3 py-2 mb-4">No missing values within shared picklists.</p>'}

    <!-- Value Diffs -->
    {'<details class="mb-4 border border-yellow-200 rounded-xl overflow-hidden"><summary class="bg-yellow-50 px-4 py-3 cursor-pointer font-semibold text-sm text-yellow-800 hover:bg-yellow-100 flex items-center justify-between"><span>Field-Level Differences in Shared Values</span><span class="text-xs font-normal">' + str(len(value_diffs)) + " values across " + str(len(vd_by_pl)) + ' picklists</span></summary><div class="p-4">' + vd_sections + "</div></details>" if value_diffs else '<p class="text-sm text-green-700 bg-green-50 rounded-lg px-3 py-2 mb-4">No field-level differences in shared values.</p>'}
  </div>

</div>
<script>
function filterContent(q) {{
  q = q.toLowerCase();
  document.querySelectorAll("details").forEach(d => {{
    d.style.display = d.innerText.toLowerCase().includes(q) ? "" : "none";
  }});
  document.querySelectorAll("tbody tr").forEach(r => {{
    r.style.display = r.innerText.toLowerCase().includes(q) ? "" : "none";
  }});
}}
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# N-tenant matrix reporters
# ---------------------------------------------------------------------------


def generate_matrix_excel_report(
    aliases: list[str],
    result: dict,
    instances: list[dict],
) -> Path:
    """
    Generate an Excel workbook for a matrix comparison of N instances.

    Sheets: Summary, Entity Coverage, Field Diffs, Field All, Picklist Diffs.
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    alias_slug = "__".join(aliases)
    fname = REPORTS_DIR / f"matrix_{alias_slug}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_matrix_summary_sheet(ws_sum, aliases, result, instances, timestamp)

    # --- Entity Coverage sheet ---
    ws_ent = wb.create_sheet("Entity Coverage")
    _write_matrix_entity_sheet(ws_ent, aliases, result)

    # --- Field Diffs sheet (non-uniform only) ---
    ws_fd = wb.create_sheet("Field Diffs")
    _write_matrix_field_sheet(ws_fd, aliases, result, diffs_only=True)

    # --- Field All sheet ---
    ws_fa = wb.create_sheet("Field All")
    _write_matrix_field_sheet(ws_fa, aliases, result, diffs_only=False)

    # --- Picklist Diffs sheet ---
    ws_pl = wb.create_sheet("Picklist Diffs")
    _write_matrix_picklist_sheet(ws_pl, aliases, result)

    wb.save(fname)
    logger.info("Matrix Excel report saved: %s", fname)
    return fname


def _write_matrix_summary_sheet(ws, aliases, result, instances, timestamp):
    s = result["summary"]
    ws.append(["Matrix Comparison Report"])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append(["Generated", timestamp])
    ws.append(["Instances", ", ".join(aliases)])
    ws.append([])
    ws.append(["Metric", "Count"])
    _header_style(ws, 5, ["Metric", "Count"])
    metrics = [
        ("Total instances", s["total_instances"]),
        ("Entities in all instances", s["entities_in_all"]),
        ("Entities with gaps", s["entities_with_gaps"]),
        ("Fields uniform across all", s["fields_uniform"]),
        ("Fields with differences", s["fields_with_diffs"]),
        ("Picklist values uniform", s["picklist_values_uniform"]),
        ("Picklist values with diffs", s["picklist_values_with_diffs"]),
    ]
    for i, (label, val) in enumerate(metrics, 6):
        ws.cell(i, 1, label)
        ws.cell(i, 2, val)
    ws.append([])
    ws.append(["Instance", "Alias", "Base URL"])
    row = len(metrics) + 8
    _header_style(ws, row - 1, ["#", "Alias", "Base URL"])
    for j, inst in enumerate(instances, 1):
        ws.append([j, inst["alias"], inst["base_url"]])
    _auto_width(ws)


def _write_matrix_entity_sheet(ws, aliases, result):
    instance_ids = [inst["id"] for inst in result["instances"]]
    headers = ["Entity"] + aliases + ["Status"]
    _header_style(ws, 1, headers)
    green = PatternFill("solid", fgColor=GREEN_FILL)
    red = PatternFill("solid", fgColor=RED_FILL)
    for row_i, (entity_name, info) in enumerate(sorted(result["entity_matrix"].items()), 2):
        ws.cell(row_i, 1, entity_name)
        for col_j, iid in enumerate(instance_ids, 2):
            cell = ws.cell(row_i, col_j)
            if iid in info["present_in"]:
                cell.value = "Present"
                cell.fill = green
            else:
                cell.value = "Missing"
                cell.fill = red
        status = "OK" if not info["missing_from"] else f"Missing in {len(info['missing_from'])}"
        ws.cell(row_i, len(aliases) + 2, status)
    _auto_width(ws)


def _write_matrix_field_sheet(ws, aliases, result, diffs_only: bool):
    instance_ids = [inst["id"] for inst in result["instances"]]
    headers = ["Entity", "Field ID", "Field Label", "Attribute"] + aliases + ["Uniform"]
    _header_style(ws, 1, headers)
    row_i = 2
    yellow = PatternFill("solid", fgColor=YELLOW_FILL)
    grey = PatternFill("solid", fgColor=GREY_FILL)
    from core.comparator import FIELD_ATTRS

    for entity_name, fields in sorted(result["field_matrix"].items()):
        for fid, finfo in sorted(fields.items()):
            if diffs_only and finfo["is_uniform"]:
                continue
            for attr in FIELD_ATTRS:
                has_diff = attr in finfo["differing_attrs"]
                if diffs_only and not has_diff:
                    continue
                ws.cell(row_i, 1, entity_name)
                ws.cell(row_i, 2, fid)
                ws.cell(row_i, 3, finfo["field_label"])
                ws.cell(row_i, 4, attr)
                for col_j, iid in enumerate(instance_ids, 5):
                    v = finfo["values"].get(iid)
                    cell = ws.cell(row_i, col_j)
                    if v is None:
                        cell.value = "(absent)"
                        cell.fill = PatternFill("solid", fgColor="FFCCCC")
                    else:
                        cell.value = v.get(attr, "")
                        if has_diff:
                            cell.fill = yellow
                        elif not diffs_only:
                            cell.fill = grey
                uniform_cell = ws.cell(row_i, len(aliases) + 5)
                uniform_cell.value = "Yes" if finfo["is_uniform"] else "No"
                row_i += 1
    _auto_width(ws)


def _write_matrix_picklist_sheet(ws, aliases, result):
    instance_ids = [inst["id"] for inst in result["instances"]]
    headers = ["Picklist ID", "Code", "Attribute"] + aliases + ["Uniform"]
    _header_style(ws, 1, headers)
    row_i = 2
    yellow = PatternFill("solid", fgColor=YELLOW_FILL)
    for pl_id, codes in sorted(result["picklist_matrix"].items()):
        for code, entry in sorted(codes.items()):
            if entry.get("is_uniform"):
                continue
            for attr in ("label_en", "status"):
                if attr not in entry:
                    continue
                ws.cell(row_i, 1, pl_id)
                ws.cell(row_i, 2, code)
                ws.cell(row_i, 3, attr)
                for col_j, iid in enumerate(instance_ids, 4):
                    val = entry[attr].get(iid)
                    cell = ws.cell(row_i, col_j)
                    cell.value = val if val is not None else "(absent)"
                    if val is None or not entry.get("is_uniform"):
                        cell.fill = yellow
                ws.cell(row_i, len(aliases) + 4, "No")
                row_i += 1
    _auto_width(ws)


def generate_matrix_html_report(
    aliases: list[str],
    result: dict,
    instances: list[dict],
    download_url: str = "",
    nav_urls: dict | None = None,
) -> str:
    """
    Generate a self-contained HTML matrix comparison report for N instances.

    Returns HTML string. Caller writes it to disk (same pattern as
    generate_html_report).
    """

    nav_urls = nav_urls or {}
    s = result["summary"]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    instance_ids = [inst["id"] for inst in instances]

    # Instance alias chips
    alias_chips = "".join(
        f'<span style="background:#e0e7ff;color:#3730a3;border-radius:12px;'
        f'padding:3px 10px;font-size:0.82rem;font-weight:600;margin-right:4px">'
        f"{escape(a)}</span>"
        for a in aliases
    )

    # Summary cards
    cards_html = "".join(
        f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;'
        f'padding:12px 18px;text-align:center;min-width:110px">'
        f'<div style="font-size:1.5rem;font-weight:700;color:{color}">{val}</div>'
        f'<div style="font-size:0.75rem;color:#64748b;margin-top:2px">{label}</div></div>'
        for label, val, color in [
            ("Instances", s["total_instances"], "#3730a3"),
            (
                "Entity Gaps",
                s["entities_with_gaps"],
                "#dc2626" if s["entities_with_gaps"] else "#16a34a",
            ),
            (
                "Field Diffs",
                s["fields_with_diffs"],
                "#d97706" if s["fields_with_diffs"] else "#16a34a",
            ),
            (
                "Picklist Diffs",
                s["picklist_values_with_diffs"],
                "#d97706" if s["picklist_values_with_diffs"] else "#16a34a",
            ),
        ]
    )

    # Field matrix table
    field_rows = []
    for entity_name, fields in sorted(result["field_matrix"].items()):
        for fid, finfo in sorted(fields.items()):
            if finfo["is_uniform"]:
                continue
            for attr in finfo["differing_attrs"]:
                cells = f'<td style="background:#f1f5f9;font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(entity_name)}</td>'
                cells += f'<td style="font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(fid)}</td>'
                cells += f'<td style="font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(attr)}</td>'
                vals_in_row = []
                for iid in instance_ids:
                    v = finfo["values"].get(iid)
                    if v is None:
                        cells += '<td style="background:#fee2e2;font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0;color:#991b1b;font-style:italic">(absent)</td>'
                        vals_in_row.append(None)
                    else:
                        val_str = escape(str(v.get(attr, "")))
                        cells += f'<td style="background:#fef9c3;font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{val_str}</td>'
                        vals_in_row.append(v.get(attr, ""))
                field_rows.append(f"<tr>{cells}</tr>")

    field_col_headers = "".join(
        f'<th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem;white-space:nowrap">{escape(a)}</th>'
        for a in aliases
    )
    field_table_html = f"""
    <table style="border-collapse:collapse;width:100%;font-size:0.85rem">
      <thead>
        <tr>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Entity</th>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Field ID</th>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Attribute</th>
          {field_col_headers}
        </tr>
      </thead>
      <tbody>
        {"".join(field_rows) if field_rows else '<tr><td colspan="100%" style="text-align:center;padding:20px;color:#16a34a">All fields uniform across instances</td></tr>'}
      </tbody>
    </table>"""

    # Entity coverage table
    ent_rows = []
    for entity_name, info in sorted(result["entity_matrix"].items()):
        cells = f'<td style="font-size:0.8rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(entity_name)}</td>'
        for iid in instance_ids:
            if iid in info["present_in"]:
                cells += '<td style="background:#dcfce7;color:#166534;text-align:center;font-size:0.75rem;padding:6px 8px;border:1px solid #e2e8f0">Present</td>'
            else:
                cells += '<td style="background:#fee2e2;color:#991b1b;text-align:center;font-size:0.75rem;padding:6px 8px;border:1px solid #e2e8f0">Missing</td>'
        ent_rows.append(f"<tr>{cells}</tr>")

    ent_col_headers = "".join(
        f'<th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">{escape(a)}</th>'
        for a in aliases
    )
    entity_table_html = f"""
    <table style="border-collapse:collapse;width:100%">
      <thead>
        <tr>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Entity</th>
          {ent_col_headers}
        </tr>
      </thead>
      <tbody>{"".join(ent_rows)}</tbody>
    </table>"""

    # Picklist diffs table
    pl_rows = []
    for pl_id, codes in sorted(result["picklist_matrix"].items()):
        for code, entry in sorted(codes.items()):
            if entry.get("is_uniform"):
                continue
            for attr in ("label_en", "status"):
                if attr not in entry:
                    continue
                cells = f'<td style="font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(pl_id)}</td>'
                cells += f'<td style="font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(str(code))}</td>'
                cells += f'<td style="font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(attr)}</td>'
                for iid in instance_ids:
                    val = entry[attr].get(iid)
                    if val is None:
                        cells += '<td style="background:#fee2e2;font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0;color:#991b1b;font-style:italic">(absent)</td>'
                    else:
                        cells += f'<td style="background:#fef9c3;font-size:0.78rem;padding:6px 8px;border:1px solid #e2e8f0">{escape(str(val))}</td>'
                pl_rows.append(f"<tr>{cells}</tr>")

    pl_col_headers = "".join(
        f'<th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">{escape(a)}</th>'
        for a in aliases
    )
    picklist_table_html = f"""
    <table style="border-collapse:collapse;width:100%">
      <thead>
        <tr>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Picklist ID</th>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Code</th>
          <th style="background:#1e3a5f;color:#fff;padding:8px 10px;font-size:0.78rem">Attribute</th>
          {pl_col_headers}
        </tr>
      </thead>
      <tbody>
        {"".join(pl_rows) if pl_rows else '<tr><td colspan="100%" style="text-align:center;padding:20px;color:#16a34a">All picklist values uniform across instances</td></tr>'}
      </tbody>
    </table>"""

    dl_link = (
        f'<a href="{escape(download_url)}" style="background:#1e3a5f;color:#fff;'
        f'border-radius:6px;padding:6px 14px;font-size:0.82rem;text-decoration:none">Download Excel</a>'
        if download_url
        else ""
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Matrix Report — {escape(", ".join(aliases))}</title>
<style>
  body{{font-family:system-ui,sans-serif;margin:0;background:#f8fafc;color:#1e293b}}
  .header{{background:#1e3a5f;color:#fff;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
  .tabs{{display:flex;gap:0;border-bottom:2px solid #e2e8f0;margin-bottom:16px}}
  .tab{{padding:8px 20px;cursor:pointer;font-size:0.85rem;font-weight:500;color:#64748b;border-bottom:2px solid transparent;margin-bottom:-2px}}
  .tab.active{{color:#1e3a5f;border-bottom-color:#1e3a5f;font-weight:700}}
  .panel{{display:none}}.panel.active{{display:block}}
  .cards{{display:flex;gap:12px;flex-wrap:wrap;margin:16px 0}}
  input[type=text]{{padding:6px 12px;border:1px solid #e2e8f0;border-radius:6px;font-size:0.85rem;width:260px}}
</style>
</head>
<body>
<div class="header">
  <div>
    <div style="font-size:1.1rem;font-weight:700">SF Config Matrix</div>
    <div style="font-size:0.82rem;opacity:0.75;margin-top:2px">{ts} &nbsp;|&nbsp; {alias_chips}</div>
  </div>
  <div>{dl_link}</div>
</div>
<div style="padding:20px 24px">
  <div class="cards">{cards_html}</div>
  <div style="margin-bottom:12px">
    <input type="text" id="srch" placeholder="Filter..." oninput="filterRows(this.value)">
  </div>
  <div class="tabs">
    <div class="tab active" onclick="showTab('fields',this)">Field Diffs</div>
    <div class="tab" onclick="showTab('entities',this)">Entity Coverage</div>
    <div class="tab" onclick="showTab('picklists',this)">Picklist Diffs</div>
  </div>
  <div id="fields" class="panel active">{field_table_html}</div>
  <div id="entities" class="panel">{entity_table_html}</div>
  <div id="picklists" class="panel">{picklist_table_html}</div>
</div>
<script>
function showTab(id,el){{
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  el.classList.add('active');
}}
function filterRows(q){{
  q=q.toLowerCase();
  document.querySelectorAll('tbody tr').forEach(r=>{{
    r.style.display=r.innerText.toLowerCase().includes(q)?'':'none';
  }});
}}
</script>
</body>
</html>"""
