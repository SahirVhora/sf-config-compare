"""Tests for N-tenant matrix comparison: comparator, reporters, routes, API."""

import json

import pytest
from core.comparator import compare_instances_matrix
from core.db import get_conn, upsert_instance
from core.reporter import generate_matrix_excel_report, generate_matrix_html_report

# ---------------------------------------------------------------------------
# Fixture: 3-instance setup (DEV, STAGING, PROD)
# ---------------------------------------------------------------------------


@pytest.fixture
def three_instances(tmp_path):
    """
    Create DEV / STAGING / PROD instances with:
    - SharedEntity in all 3 (field1 uniform, field_C differs STAGING vs others)
    - OnlyInDevProd entity missing from STAGING
    - Picklist PL1 with different label_en per instance
    """
    ids = []
    for alias, url in [
        ("DEV", "https://dev.example.com"),
        ("STAGING", "https://staging.example.com"),
        ("PROD", "https://prod.example.com"),
    ]:
        iid = upsert_instance(
            {
                "alias": alias,
                "base_url": url,
                "company_id": f"{alias}001",
                "auth_type": "basic",
                "username": "admin",
                "client_id": None,
                "token_url": None,
            }
        )
        ids.append(iid)

    dev_id, staging_id, prod_id = ids

    with get_conn() as conn:
        # SharedEntity in all 3
        for inst_id in ids:
            conn.execute(
                "INSERT INTO metadata_entities "
                "(instance_id, entity_name, entity_label, element_name, pull_timestamp)"
                " VALUES (?, ?, ?, ?, datetime('now'))",
                (inst_id, "SharedEntity", "Shared Entity", "SharedEntity"),
            )

        # OnlyInDevProd only in DEV and PROD (not STAGING)
        for inst_id in [dev_id, prod_id]:
            conn.execute(
                "INSERT INTO metadata_entities "
                "(instance_id, entity_name, entity_label, element_name, pull_timestamp)"
                " VALUES (?, ?, ?, ?, datetime('now'))",
                (inst_id, "OnlyInDevProd", "Only DevProd", "OnlyInDevProd"),
            )

        # Fields for SharedEntity
        rows = conn.execute(
            "SELECT id, instance_id FROM metadata_entities WHERE entity_name='SharedEntity'",
        ).fetchall()
        shared_eid_map = {r["instance_id"]: r["id"] for r in rows}

        # field1 is uniform across all 3
        for inst_id in ids:
            eid = shared_eid_map[inst_id]
            conn.execute(
                "INSERT INTO metadata_fields "
                "(entity_id, field_id, field_label, field_type, required, visibility,"
                " max_length, picklist_id, is_custom, raw_attributes)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    eid,
                    "field1",
                    "Field 1",
                    "Edm.String",
                    "false",
                    "true",
                    "255",
                    "",
                    0,
                    None,
                ),
            )

        # field_C differs: STAGING has required=true, others false
        for inst_id, req_val in [
            (dev_id, "false"),
            (staging_id, "true"),
            (prod_id, "false"),
        ]:
            eid = shared_eid_map[inst_id]
            conn.execute(
                "INSERT INTO metadata_fields "
                "(entity_id, field_id, field_label, field_type, required, visibility,"
                " max_length, picklist_id, is_custom, raw_attributes)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    eid,
                    "field_C",
                    "Field C",
                    "Edm.String",
                    req_val,
                    "true",
                    "255",
                    "",
                    0,
                    None,
                ),
            )

        # Picklist PL1 with different label_en per instance
        for inst_id, label in [
            (dev_id, "Label DEV"),
            (staging_id, "Label STG"),
            (prod_id, "Label PROD"),
        ]:
            conn.execute(
                "INSERT INTO picklist_values "
                "(instance_id, picklist_id, option_id, external_code, parent_picklist_id,"
                " status, label_en, all_labels, pull_timestamp)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))",
                (
                    inst_id,
                    "status",
                    "OPT1",
                    "PL1",
                    None,
                    "ACTIVE",
                    label,
                    json.dumps({"en_US": label}),
                ),
            )

    return dev_id, staging_id, prod_id


# ---------------------------------------------------------------------------
# comparator tests
# ---------------------------------------------------------------------------


def test_matrix_result_structure(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    for key in (
        "instances",
        "summary",
        "entity_matrix",
        "field_matrix",
        "picklist_matrix",
    ):
        assert key in result, f"Missing top-level key: {key}"

    s = result["summary"]
    for skey in (
        "total_instances",
        "entities_in_all",
        "entities_with_gaps",
        "fields_uniform",
        "fields_with_diffs",
        "picklist_values_uniform",
        "picklist_values_with_diffs",
    ):
        assert skey in s, f"Missing summary key: {skey}"

    assert s["total_instances"] == 3


def test_matrix_field1_is_uniform(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    shared_fields = result["field_matrix"].get("SharedEntity", {})
    assert "field1" in shared_fields
    assert shared_fields["field1"]["is_uniform"] is True
    assert shared_fields["field1"]["differing_attrs"] == []


def test_matrix_field_c_is_not_uniform(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    shared_fields = result["field_matrix"].get("SharedEntity", {})
    assert "field_C" in shared_fields
    assert shared_fields["field_C"]["is_uniform"] is False
    assert "required" in shared_fields["field_C"]["differing_attrs"]


def test_matrix_entity_coverage_gap(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    em = result["entity_matrix"]
    assert "OnlyInDevProd" in em
    info = em["OnlyInDevProd"]
    assert staging_id in info["missing_from"]
    assert dev_id in info["present_in"]
    assert prod_id in info["present_in"]


def test_matrix_entities_with_gaps_count(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])
    assert result["summary"]["entities_with_gaps"] >= 1


def test_matrix_picklist_diffs(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    pl = result["picklist_matrix"]
    assert "status" in pl
    assert "PL1" in pl["status"]
    entry = pl["status"]["PL1"]
    assert entry["is_uniform"] is False
    assert entry["label_en"][dev_id] == "Label DEV"
    assert entry["label_en"][staging_id] == "Label STG"
    assert entry["label_en"][prod_id] == "Label PROD"


def test_matrix_requires_min_two_instances():
    with pytest.raises(ValueError, match="at least 2"):
        compare_instances_matrix([1])


def test_matrix_summary_uniform_count(three_instances):
    dev_id, staging_id, prod_id = three_instances
    result = compare_instances_matrix([dev_id, staging_id, prod_id])
    # field1 is uniform across all 3
    assert result["summary"]["fields_uniform"] >= 1
    # field_C differs
    assert result["summary"]["fields_with_diffs"] >= 1


# ---------------------------------------------------------------------------
# reporter tests
# ---------------------------------------------------------------------------


def test_matrix_excel_generates_file(three_instances, tmp_path, monkeypatch):
    import config

    monkeypatch.setattr(config, "REPORTS_DIR", tmp_path)

    dev_id, staging_id, prod_id = three_instances
    from core.db import get_instance

    instances = [get_instance(i) for i in [dev_id, staging_id, prod_id]]
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    path = generate_matrix_excel_report(["DEV", "STAGING", "PROD"], result, instances)
    assert path.exists()
    assert path.suffix == ".xlsx"

    import openpyxl

    wb = openpyxl.load_workbook(str(path))
    assert "Summary" in wb.sheetnames
    assert "Entity Coverage" in wb.sheetnames
    assert "Field Diffs" in wb.sheetnames
    assert "Field All" in wb.sheetnames
    assert "Picklist Diffs" in wb.sheetnames


def test_matrix_html_contains_instance_aliases(three_instances):
    dev_id, staging_id, prod_id = three_instances
    from core.db import get_instance

    instances = [get_instance(i) for i in [dev_id, staging_id, prod_id]]
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    html = generate_matrix_html_report(["DEV", "STAGING", "PROD"], result, instances)
    assert "DEV" in html
    assert "STAGING" in html
    assert "PROD" in html
    assert "<!DOCTYPE html>" in html


def test_matrix_html_shows_field_diffs(three_instances):
    dev_id, staging_id, prod_id = three_instances
    from core.db import get_instance

    instances = [get_instance(i) for i in [dev_id, staging_id, prod_id]]
    result = compare_instances_matrix([dev_id, staging_id, prod_id])

    html = generate_matrix_html_report(["DEV", "STAGING", "PROD"], result, instances)
    assert "field_C" in html


# ---------------------------------------------------------------------------
# Flask route tests
# ---------------------------------------------------------------------------


def test_matrix_route_get(client):
    resp = client.get("/matrix")
    assert resp.status_code == 200
    assert b"Matrix" in resp.data


def test_matrix_route_post_one_instance_flashes_error(client, three_instances):
    dev_id, _, _ = three_instances
    resp = client.post(
        "/matrix",
        data={"instance_ids[]": [str(dev_id)], "csrf_token": "test"},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert b"at least 2" in resp.data


def test_matrix_route_post_runs_comparison(client, three_instances):
    dev_id, staging_id, prod_id = three_instances
    resp = client.post(
        "/matrix",
        data={
            "instance_ids[]": [str(dev_id), str(staging_id), str(prod_id)],
            "compare_fields": ["label_en", "status"],
            "csrf_token": "test",
        },
        follow_redirects=False,
    )
    # Should redirect to report view
    assert resp.status_code == 302
    assert "/reports/" in resp.headers.get("Location", "")


# ---------------------------------------------------------------------------
# API endpoint tests
# ---------------------------------------------------------------------------


def test_matrix_api_returns_200(client, three_instances):
    dev_id, staging_id, prod_id = three_instances
    resp = client.post(
        "/api/v1/matrix",
        json={"instance_ids": [dev_id, staging_id, prod_id]},
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert "summary" in data
    assert "entity_matrix" in data
    assert "field_matrix" in data
    assert "picklist_matrix" in data
    assert data["summary"]["total_instances"] == 3


def test_matrix_api_rejects_single_instance(client, three_instances):
    dev_id, _, _ = three_instances
    resp = client.post(
        "/api/v1/matrix",
        json={"instance_ids": [dev_id]},
        content_type="application/json",
    )
    assert resp.status_code == 422


def test_matrix_api_rejects_empty_ids(client):
    resp = client.post(
        "/api/v1/matrix",
        json={"instance_ids": []},
        content_type="application/json",
    )
    assert resp.status_code == 422


def test_matrix_api_rejects_unknown_instance(client):
    resp = client.post(
        "/api/v1/matrix",
        json={"instance_ids": [9999, 9998]},
        content_type="application/json",
    )
    assert resp.status_code == 404


def test_matrix_api_include_reports_generates_file(
    client, three_instances, tmp_path, monkeypatch
):
    import config

    monkeypatch.setattr(config, "REPORTS_DIR", tmp_path)
    import core.reporter as _rep

    monkeypatch.setattr(_rep, "REPORTS_DIR", tmp_path)

    dev_id, staging_id, prod_id = three_instances
    resp = client.post(
        "/api/v1/matrix",
        json={
            "instance_ids": [dev_id, staging_id, prod_id],
            "include_reports": True,
        },
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert "report_id" in data
